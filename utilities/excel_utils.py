import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
import numpy as np


def excel_column_to_list(file_path, column_name):
    df = pd.read_excel(file_path)
    return df[column_name].tolist()


def save_results_to_excel(results, excel_path):
    try:
        df = pd.read_excel(excel_path)

        # 检查Q和R列是否存在，如果存在则清除除标题外的所有数据
        if '代码' in df.columns and '价格' in df.columns:
            df.loc[0:, '代码'] = pd.NA  # 清除Q列除了第一行外的所有数据
            df.loc[0:, '价格'] = pd.NA  # 清除R列除了第一行外的所有数据
    except FileNotFoundError:
        # 当文件不存在时，创建一个新的DataFrame
        df = pd.DataFrame(columns=['代码', '价格'])

    for idx, (code, price) in enumerate(results, start=1):
        price = round(float(price), 1)
        df.at[idx-1, '代码'] = code
        df.at[idx-1, '价格'] = price

    df.to_excel(excel_path, index=False)


def find_data(df, row_name, column_name):
    rows = row_name.split('+')
    columns = column_name.split('+')

    data_sum = 0
    data_found = False  # 添加一个标志变量来记录是否找到数据

    for index, row in df.iterrows():
        row_parts = index.split('+')
        if any(r in row_parts for r in rows):
            for df_col in df.columns:
                col_parts = df_col.split('+')
                if any(col in col_parts for col in columns) and pd.notna(row[df_col]):
                    value = row[df_col]
                    if isinstance(value, (int, float)):
                        data_sum += value
                        data_found = True  # 设置标志为True表示找到了数据
                    else:
                        print(f"Non-numeric value found: {value}, Row: {index}, Column: {df_col}")

    # 如果没有找到数据，返回None
    if not data_found:
        return None

    # 如果找到了数据，按照之前的逻辑处理
    if data_sum % 1 == 0:  # 如果是整数
        return int(data_sum)  # 返回整数形式
    else:  # 如果是带小数的数值
        return round(data_sum, 2)  # 保留两位小数


def split_excel_by_blank_rows(file_path, sheet_name):
    # Excel列名从A开始，因此"Q"列是第17列，但是在 pandas 中列索引从0开始计数
    start_col_index = 16  # "Q"列的索引

    # 读取从"Q"列开始的所有列
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, index_col=None,
                       usecols=lambda x: x >= start_col_index)
    print(df.shape)
    # 查找空白行的索引以横向分割
    blank_row_indices = df[df.isnull().all(axis=1)].index

    dataframes = []
    start_row = 0
    for end_row in blank_row_indices:
        if end_row - start_row >= 2:  # 空白行超过2行，分割数据块
            block = df.iloc[start_row:end_row]
            # 纵向分割：删除全部为空的行
            block = block.dropna(axis=0, how='all')
            if block.shape[0] > 0 and block.shape[1] > 0:
                # 设置第一行和第一列为列名和行名
                block.columns = block.iloc[0]
                block = block[1:]
                block = block.set_index(block.columns[0])
                dataframes.append(block)
            start_row = end_row + 1
    print(f'start_row: {start_row}, len of df: {len(df)}')
    # 添加最后一个数据块
    if start_row < len(df):
        # 查找最后一个数据块的真实起始行
        last_block_start = df[start_row:].dropna(how='all').index[0]
        block = df.iloc[last_block_start:]
        block = block.dropna(axis=1, how='all')
        if block.shape[0] > 0 and block.shape[1] > 0:
            block.columns = block.iloc[0]
            block = block[1:]
            block = block.set_index(block.columns[0])
            dataframes.append(block)

    return dataframes


def process_data_blocks(dataframes):
    processed_blocks = []
    for df in dataframes:
        if df.shape[0] > 0 and df.shape[1] > 0:
            # 检查并删除第一行如果它是列标题的重复
            if df.iloc[0, 0] == '形态':
                df = df.iloc[1:]

            # 移除以 "Unnamed" 开头的列
            # df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

            # 移除完全为空的列
            df = df.dropna(axis=1, how='all')

            # 将“形态”列作为行名（如果存在）
            if '形态' in df.columns:
                df = df.set_index('形态')

            # 可以在这里添加更多针对每个数据块的处理逻辑

            processed_blocks.append(df)

    return processed_blocks


# def save_new_dfs_to_excel(new_dfs, conclusions, file_path):
#     try:
#         df = pd.read_excel(file_path)
#         # 检查Q和R列是否存在，如果存在则清除除标题外的所有数据
#         if '历史' in df.columns :
#             df.loc[0:, '代码'] = pd.NA  # 清除Q列除了第一行外的所有数据
#
#         startrow = 1
#         startcol = 20  # "U"列是第21列，但索引从0开始，所以用20
#
#         # 循环处理每个新的DataFrame
#         for new_df in new_dfs:
#             # 直接写入新数据，从"U"列开始
#             new_df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, startcol=startcol, index=True)
#
#             # 更新起始行位置
#             startrow += len(new_df.index) + 6  # 为下一个数据块留出空间
#
#         # 循环处理结论里的DataFrame
#         conclusions[0].to_excel(writer, sheet_name='Sheet1', startrow=0, startcol=30, index=True)
#         conclusions[1].to_excel(writer, sheet_name='Sheet1', startrow=0, startcol=32, index=True)
#
#     df.to_excel(file_path, index=False)
#     with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#         startrow = 1
#         startcol = 20  # "U"列是第21列，但索引从0开始，所以用20
#
#         # 循环处理每个新的DataFrame
#         for new_df in new_dfs:
#             # 直接写入新数据，从"U"列开始
#             new_df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, startcol=startcol, index=True)
#
#             # 更新起始行位置
#             startrow += len(new_df.index) + 6  # 为下一个数据块留出空间
#
#         # 循环处理结论里的DataFrame
#         conclusions[0].to_excel(writer, sheet_name='Sheet1', startrow=0, startcol=30, index=True)
#         conclusions[1].to_excel(writer, sheet_name='Sheet1', startrow=0, startcol=32, index=True)


def save_new_dfs_to_excel(new_dfs, conclusions, file_path):
    # 尝试读取现有数据
    try:
        df = pd.read_excel(file_path)

        # 清除指定列的数据
        for col_index in range(20, 35):  # 从第21列到第35列（"U"到"AI"）
            if col_index < len(df.columns):
                df.iloc[0:, col_index] = pd.NA  # 从第二行开始清除数据

    except FileNotFoundError:
        # 当文件不存在时，创建一个新的DataFrame
        df = pd.DataFrame()

    # 保存清除后的DataFrame
    df.to_excel(file_path, index=False)

    # 写入新数据
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        startrow = 0
        startcol = 20  # "U"列

        # 循环处理每个新的DataFrame
        for new_df in new_dfs:
            new_df.to_excel(writer, sheet_name='Sheet1', startrow=startrow, startcol=startcol, index=True)
            startrow += len(new_df.index) + 6  # 为下一个数据块留出空间

        # 写入结论里的DataFrame
        conclusions.to_excel(writer, sheet_name='Sheet1', startrow=0, startcol=30, index=True)
        # conclusions[0].to_excel(writer, sheet_name='Sheet1', startrow=0, startcol=30, index=True)
        # conclusions[1].to_excel(writer, sheet_name='Sheet1', startrow=0, startcol=32, index=True)


def load_workbooks(input_path, output_path):
    wb_input = load_workbook(filename=input_path, data_only=True)
    ws_input = wb_input['Sheet1']

    wb_output = load_workbook(filename=output_path, data_only=True)
    ws_output = wb_output['Sheet1']

    return wb_input, ws_input, wb_output, ws_output


def sync_and_clear_extra_data(ws_input, ws_output, columns, relation_dict):
    for column in columns:
        row, input_row_count = 3, 0
        while ws_input[f'{column}{row}'].value is not None:
            ws_output[f'{column}{row}'].value = ws_input[f'{column}{row}'].value
            row += 1
            input_row_count += 1

        while ws_output[f'{column}{row}'].value is not None:
            ws_output[f'{column}{row}'].value = None
            row += 1

        related_column = relation_dict[column]
        output_related_row = 3 + input_row_count
        while ws_output[f'{related_column}{output_related_row}'].value is not None:
            ws_output[f'{related_column}{output_related_row}'].value = None
            output_related_row += 1


def save_workbooks(wb_input, wb_output, input_path, output_path):
    wb_output.save(output_path)
    wb_input.save(input_path)


def create_dataframe_from_columns(ws, columns, relation_dict):
    data = []
    for column in columns:
        row = 3
        while ws[f'{column}{row}'].value is not None:
            code = ws[f'{column}{row}'].value
            data.append({'code': code, 'position': f'{relation_dict[column]}{row}'})
            row += 1
    return pd.DataFrame(data)


