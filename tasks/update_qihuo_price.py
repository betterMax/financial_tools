from openpyxl import load_workbook
import pandas as pd
from utilities.web_scraping_utils import get_latest_price


def run(path, mode):
    # 加载工作簿
    wb = load_workbook(filename=path, data_only=True)
    ws = wb['Sheet1']

    # 待处理的sheet列表
    relation_dict = {'A': 'D', 'B': 'E', 'C': 'F'}
    if mode == 'work':
        columns = ['A', 'B']
    else:
        columns = ['A', 'B', 'C']

    # 创建DataFrame
    data = []
    for column in columns:
        row = 3
        while ws[f'{column}{row}'].value is not None:
            code = ws[f'{column}{row}'].value
            data.append({'code': code, 'position': f'{relation_dict[column]}{row}', 'price': 0})
            row += 1

    df = pd.DataFrame(data)

    # 去重并获取价格
    unique_codes = df['code'].unique()
    for code in unique_codes:
        price = get_latest_price(code)  # 假设这个函数返回价格
        df.loc[df['code'] == code, 'price'] = price

    # 更新回工作簿
    for _, row in df.iterrows():
        ws[row['position']] = row['price']

    # 保存工作簿
    wb.save(path)
