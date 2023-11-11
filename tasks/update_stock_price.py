from openpyxl import load_workbook
import time
from utilities.web_scraping_utils import init_chrome, get_stock_latest_price


def run(path, mode):
    # 加载工作簿
    wb = load_workbook(filename=path, data_only=True)
    driver = init_chrome()

    # 待处理的sheet列表
    relation_dict = {'A':'D', 'B':'E', 'C':'F'}
    if mode == 'work':
        columns = ['A', 'B']
    else:
        columns = ['A', 'B', 'C']

    for column in columns:
        # 选择工作表
        ws = wb['Sheet1']

        # 初始化行号
        row = 3

        # 循环处理每一行，直到A列没有数据
        while ws[f'{column}{row}'].value is not None:
            # 读取单元格
            code = ws[f'{column}{row}'].value
            print(f'update {code} price')
            latest_price = get_stock_latest_price(driver, code)
            price = None
            # 尝试将价格转换为浮点数
            try:
                if latest_price is not None:
                    price = float(latest_price)
                else:
                    print(f'Warning: Latest price is None. Please check the price.')
            except ValueError:
                print(f'Warning: Cannot convert price "{latest_price}" to number. Please check the price.')
            else:
                ws[f'{relation_dict[column]}{row}'].value = price

            # 处理下一行
            row += 1

            # 暂停一段时间
            time.sleep(1)  # 这里设置暂停一秒，可以根据实际需要调整
    driver.quit()
    # 保存修改
    wb.save(path)