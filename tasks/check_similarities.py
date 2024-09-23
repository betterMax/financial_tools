import pandas as pd
from difflib import SequenceMatcher
import openpyxl
from collections import defaultdict
from utilities.similarity_utils import SimilarityCalculator


def calculate_similarity(str1, str2):
    return SequenceMatcher(None, str1, str2).ratio()


def process_operation_reason(reason):
    if len(reason) == 3:
        if "支撑" in reason:
            return ["支撑位真跌破", "支撑位假跌破"]
        elif "压力" in reason:
            return ["压力位真跌破", "压力位假跌破"]
    elif "支撑位" in reason:
        return ["趋势真跌破", "趋势假跌破"]
    elif "压力位" in reason:
        return ["趋势真突破", "趋势假突破"]
    else:
        return None


def load_data_source_a(file_path):
    try:
        df_a = pd.read_excel(file_path, sheet_name=0, usecols="P:T", header=1)
        return df_a
    except Exception as e:
        print(f"加载数据源A时发生错误: {e}")
        return None


def load_data_source_b(file_path):
    try:
        df_b = pd.read_excel(file_path, sheet_name=0)
        return df_b
    except Exception as e:
        print(f"加载数据源B时发生错误: {e}")
        return None


def calculate_similarity_for_top_n(similarity_list, top_n):
    sorted_similarities = sorted(similarity_list, reverse=True)
    top_n_similarities = sorted_similarities[:top_n]
    avg_similarity = sum(top_n_similarities) / top_n
    return avg_similarity


def check_operation_type_count(df, threshold=10):
    operation_reason_counts = df['操作原因'].value_counts()
    for operation_reason, count in operation_reason_counts.items():
        if count < threshold:
            print(f"操作原因 '{operation_reason}' 的行数少于 {threshold}，已移除。")
            df = df[df['操作原因'] != operation_reason]
    return df


def process_and_match(file_path_a, file_path_b, top_3=3, top_5=5):
    try:
        df_a = load_data_source_a(file_path_a)
        df_a = df_a.dropna()
        df_b = load_data_source_b(file_path_b)
        df_b = check_operation_type_count(df_b)
        df_b = df_b.dropna()

        if df_a is None or df_b is None:
            return

        similarity_calculator = SimilarityCalculator(method="jaccard")
        results = []

        for index, row_a in df_a.iterrows():
            print(f"开始处理行 {index}，标的 {row_a['候选标的代码']}的数据")
            similarity_list = []

            if row_a['形态'] == 0 or row_a['中期趋势'] == 0 or row_a['长期趋势'] == 0:
                print(f"第 {index} 行包含 0 值，跳过此行并将相似度设为 0: {row_a}")
                similarity_list.append({
                    "操作原因B": None,
                    "长期趋势": row_a['长期趋势'],
                    "中期趋势": row_a['中期趋势'],
                    "形态": row_a['形态'],
                    "相似度": 0
                })
                continue

            operation_reason_a = row_a['操作原因']
            match_criteria = process_operation_reason(operation_reason_a)

            if not match_criteria:
                print(f"跳过无效的操作原因: {operation_reason_a}")
                continue

            matches_in_b = df_b[df_b['操作原因'].isin(match_criteria)]

            for _, row_b in matches_in_b.iterrows():
                try:
                    similarity_result = similarity_calculator.calculate_total_similarity(
                        row_a['长期趋势'], row_b['长期趋势'],
                        row_a['中期趋势'], row_b['中期趋势'],
                        row_a['形态'], row_b['形态']
                    )

                    similarity_list.append({
                        "操作原因B": row_b['操作原因'],
                        "长期趋势": row_b['长期趋势'],
                        "中期趋势": row_b['中期趋势'],
                        "形态": row_b['形态'],
                        "长期趋势相似度": similarity_result['similarity_long'],
                        "中期趋势相似度": similarity_result['similarity_middle'],
                        "形态相似度": similarity_result['similarity_shape'],
                        "总相似度": similarity_result['total_similarity']
                    })

                except Exception as e:
                    print(f"处理行 {index} 时发生错误: {e}")

            grouped_similarity = defaultdict(list)
            for item in similarity_list:
                operation_reason_b = item['操作原因B']
                if operation_reason_b not in grouped_similarity:
                    grouped_similarity[operation_reason_b] = []
                grouped_similarity[operation_reason_b].append(item['总相似度'])

            for operation_reason_b, similarities in grouped_similarity.items():
                if len(similarities) <= 5:
                    print(
                        f"跳过操作思路: {operation_reason_b}和标的 {row_a['候选标的代码']}，因为相似度列表长度小于等于 5")
                    continue

                s_at_3 = calculate_similarity_for_top_n(similarities, top_3)
                s_at_5 = calculate_similarity_for_top_n(similarities, top_5)

                results.append({
                    "候选标的代码": row_a['候选标的代码'],
                    "操作思路": operation_reason_b,
                    "S@3": s_at_3,
                    "S@5": s_at_5
                })

        return results

    except Exception as e:
        print(f"处理数据时发生错误: {e}")
        return None


def run(input_path, output_path, future_hisotry_database_path):
    final_results = process_and_match(input_path, future_hisotry_database_path)

    if final_results:
        for result in final_results:
            print(
                f"标的: {result['候选标的代码']} - 操作思路: {result['操作思路']} - S@3: {result['S@3']:.2f}"
                f" - S@5: {result['S@5']:.2f}")
    else:
        print("没有结果可显示。")

    # 加载现有的 input, output Excel 文件
    wb_input = openpyxl.load_workbook(input_path)
    ws_input = wb_input.active
    wb_output = openpyxl.load_workbook(output_path)
    ws_output = wb_output.active

    grouped_results = defaultdict(list)
    for result in final_results:
        grouped_results[result['候选标的代码']].append(result)

    # 找到目标范围 "P2:T" 对应的行位置，假设行索引从 3 开始
    for code, result_list in grouped_results.items():
        for row in ws_input.iter_rows(min_row=3, max_col=16, values_only=False):  # P 列为第16列
            if row[15].value == code:  # 代码在第16列 (P列)
                # 针对这个代码，如果 result_list 中有1个或者2个结果，分别保存
                if len(result_list) == 1:
                    # 保存到 Q 列（第17列）
                    operation_reason_output = f"{result_list[0]['操作思路']}：S@3 = {result_list[0]['S@3']:.3f}, " \
                                              f"S@5 = {result_list[0]['S@5']:.3f}"
                    ws_input.cell(row=row[0].row, column=21, value=operation_reason_output)  # 保存到第21列 (U列)
                elif len(result_list) == 2:
                    # 保存到 Q 列（第16列）和 R 列（第17列）
                    operation_reason_output = f"{result_list[0]['操作思路']}：S@3 = {result_list[0]['S@3']:.3f}, " \
                                              f"S@5 = {result_list[0]['S@5']:.3f}; {result_list[1]['操作思路']}: " \
                                              f"S@3 = {result_list[1]['S@3']:.3f}, S@5 = {result_list[1]['S@5']:.3f}"

                    ws_input.cell(row=row[0].row, column=21, value=operation_reason_output)  # 保存到第21列 (U列)
                    # ws_input.cell(row=row[0].row, column=22, value=operation_reason_output_2)  # 保存到第22列 (V列)

    # 清空 output 文件的 "P2:U" 范围数据
    for row in ws_output.iter_rows(min_row=2, max_row=ws_output.max_row, min_col=16, max_col=21):
        for cell in row:
            cell.value = None  # 清空单元格

    # 复制 input 文件 "P2:V" 数据到 output 文件
    for row_idx, row in enumerate(
            ws_input.iter_rows(min_row=2, max_row=ws_input.max_row, min_col=16, max_col=21, values_only=True), start=2):
        # 检查 P 列是否为空，如果为空则结束复制过程
        if row[0] is None:  # 这里 row[0] 对应 P 列的单元格 (第16列)
            break

        # 将每一行的值逐个写入 output 文件
        for col_idx, value in enumerate(row, start=16):  # 从第16列 (P 列) 开始写入
            ws_output.cell(row=row_idx, column=col_idx, value=value)

    # 保存 Excel 文件
    wb_input.save(input_path)
    wb_output.save(output_path)
    print(f"结果已保存到原文件 {input_path}和输出文件 {output_path}")
